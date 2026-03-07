import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime, timedelta
from collections import Counter
import io
import pickle

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="INTERLOG · KPI Dashboard",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ─────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700;800&family=Barlow:wght@300;400;500;600&display=swap');

html, body, [class*="css"] {
    font-family: 'Barlow', sans-serif;
    background-color: #0A1628;
    color: #F0F4F8;
}
.stApp { background: #0A1628; }

h1, h2, h3 {
    font-family: 'Barlow Condensed', sans-serif;
    letter-spacing: 1px;
}

/* Cards métricas */
.metric-card {
    background: #132236;
    border: 1px solid rgba(0,201,167,0.15);
    border-top: 3px solid #00C9A7;
    border-radius: 8px;
    padding: 1.2rem;
    text-align: center;
}
.metric-value {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 2.4rem; font-weight: 800;
    color: #00C9A7; line-height: 1;
}
.metric-value.orange { color: #FF8C42; }
.metric-value.red    { color: #FF3D5E; }
.metric-value.gold   { color: #FFD060; }
.metric-label {
    font-size: 0.7rem; font-weight: 600;
    color: #6B8099; letter-spacing: 1px;
    text-transform: uppercase; margin-top: 0.3rem;
}
.metric-sub {
    font-size: 0.75rem; color: #9AB0C4; margin-top: 0.2rem;
}

/* Section headers */
.section-header {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 1.1rem; font-weight: 700;
    color: #00C9A7; letter-spacing: 2px;
    text-transform: uppercase;
    border-left: 3px solid #00C9A7;
    padding-left: 0.8rem; margin: 1.5rem 0 0.8rem;
}

/* Steps */
.step-badge {
    display: inline-block;
    background: #00C9A7; color: #0A1628;
    font-family: 'Barlow Condensed', sans-serif;
    font-weight: 800; font-size: 0.9rem;
    padding: 2px 10px; border-radius: 4px;
    margin-right: 0.5rem;
}

/* Upload zone */
.upload-card {
    background: #132236;
    border: 1px dashed rgba(0,201,167,0.3);
    border-radius: 8px; padding: 1.5rem;
    text-align: center;
}

/* Alert */
.alert-info {
    background: rgba(0,201,167,0.1);
    border: 1px solid rgba(0,201,167,0.3);
    border-radius: 6px; padding: 0.8rem 1rem;
    font-size: 0.85rem; color: #9AB0C4;
}
.alert-warn {
    background: rgba(255,140,66,0.1);
    border: 1px solid rgba(255,140,66,0.4);
    border-radius: 6px; padding: 0.8rem 1rem;
    font-size: 0.85rem; color: #FF8C42;
}
.alert-success {
    background: rgba(0,201,167,0.15);
    border: 1px solid #00C9A7;
    border-radius: 6px; padding: 0.8rem 1rem;
    font-size: 0.85rem; color: #00C9A7;
    font-weight: 600;
}

/* Tablas */
.dataframe { background: #132236 !important; }
thead tr th { background: #007A65 !important; color: white !important; }

/* Buttons */
.stButton > button {
    background: #00C9A7; color: #0A1628;
    font-family: 'Barlow Condensed', sans-serif;
    font-weight: 700; font-size: 1rem;
    letter-spacing: 1px; border: none;
    border-radius: 6px; padding: 0.6rem 2rem;
    transition: all 0.2s;
}
.stButton > button:hover { background: #00E0BA; transform: translateY(-1px); }

/* Sidebar */
.css-1d391kg { background: #0F2040; }

/* Hide streamlit branding */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────
FASA = 'FINNING ARGENTINA SOCIEDAD ANO'
FSM  = 'FINNING SOLUCIONES MINERAS SA'

LIMITES_LIB = {
    'AVION':    {'VERDE': 1, 'NARANJA': 3, 'ROJO': 4},
    'MARITIMO': {'VERDE': 3, 'NARANJA': 4, 'ROJO': 5},
    'CAMION':   {'VERDE': 1, 'NARANJA': 2, 'ROJO': 3},
}

COLORS = {
    'accent':  '#00C9A7',
    'orange':  '#FF8C42',
    'red':     '#FF3D5E',
    'gold':    '#FFD060',
    'bg':      '#132236',
    'bg2':     '#1A2E48',
    'gray':    '#6B8099',
    'verde':   '#00C9A7',
    'naranja': '#FF8C42',
    'rojo':    '#FF3D5E',
}

CHART_LAYOUT = dict(
    paper_bgcolor='rgba(0,0,0,0)',
    plot_bgcolor='rgba(19,34,54,0.8)',
    font=dict(family='Barlow, sans-serif', color='#9AB0C4'),
    margin=dict(l=10, r=10, t=30, b=10),
    xaxis=dict(gridcolor='rgba(107,128,153,0.15)', linecolor='rgba(107,128,153,0.3)'),
    yaxis=dict(gridcolor='rgba(107,128,153,0.15)', linecolor='rgba(107,128,153,0.3)'),
)

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def parse_date(val):
    if isinstance(val, datetime): return val
    if isinstance(val, str):
        for fmt in ['%d/%m/%Y %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d']:
            try: return datetime.strptime(val, fmt)
            except: pass
    return None

def dias_habiles(d1, d2):
    """Días hábiles entre d1 y d2, sin contar d1 (día de oficialización)"""
    if not d1 or not d2: return None
    d1 = d1.date() if hasattr(d1, 'date') else d1
    d2 = d2.date() if hasattr(d2, 'date') else d2
    if d2 <= d1: return 0
    dias = 0
    cur = d1 + timedelta(days=1)
    while cur <= d2:
        if cur.weekday() < 5:
            dias += 1
        cur += timedelta(days=1)
    return dias

def business_hours(d1, d2):
    if not d1 or not d2: return None
    if d2 < d1: d1, d2 = d2, d1
    total = 0
    current = d1
    while current < d2:
        if current.weekday() < 5:
            start = current.replace(hour=9, minute=0, second=0, microsecond=0)
            end   = current.replace(hour=18, minute=0, second=0, microsecond=0)
            if current < start: current = start
            day_end = min(end, d2)
            if current < day_end:
                total += (day_end - current).total_seconds() / 3600
        current = (current + timedelta(days=1)).replace(hour=9, minute=0, second=0, microsecond=0)
    return total

def limite_ofi(razon, via):
    return 48 if razon == FSM and via == 'MARITIMO' else 24

def color_kpi(pct):
    if pct >= 95: return COLORS['accent']
    if pct >= 80: return COLORS['orange']
    return COLORS['red']

def metric_html(value, label, sub=None, color='accent'):
    color_class = '' if color == 'accent' else color
    sub_html = f'<div class="metric-sub">{sub}</div>' if sub else ''
    return f"""
    <div class="metric-card">
        <div class="metric-value {color_class}">{value}</div>
        <div class="metric-label">{label}</div>
        {sub_html}
    </div>
    """

# ─────────────────────────────────────────────
# PROCESAMIENTO
# ─────────────────────────────────────────────
def procesar_liberadas(df):
    results = []
    for _, r in df.iterrows():
        razon = r.get('Razon Social', '')
        via   = str(r.get('Via', '')).upper().strip()
        canal = str(r.get('Canal', '')).upper().strip()
        f_ofi = parse_date(r.get('Fecha Oficialización'))
        f_can = parse_date(r.get('Fecha Cancelada'))
        dias   = dias_habiles(f_ofi, f_can)
        limite = LIMITES_LIB.get(via, {}).get(canal, 9999)
        desvio = dias is not None and dias > limite
        results.append({
            'razon': razon, 'nombre': 'FASA' if razon == FASA else 'FSM',
            'ref': r.get('Referencia',''), 'carpeta': r.get('Carpeta',''),
            'via': via, 'canal': canal,
            'f_ofi': f_ofi, 'f_cancel': f_can,
            'hs': dias,
            'limite': limite, 'desvio': desvio,
            'desvio_desc': '', 'parametro': ''
        })
    return results

def procesar_oficializados(df):
    results = []
    for _, r in df.iterrows():
        razon = r.get('Razon Social', '')
        via   = str(r.get('Via', '')).upper().strip()
        f_ofi = parse_date(r.get('Fecha Oficialización'))
        f_ult = parse_date(r.get('Ultimo Evento'))
        hs    = business_hours(f_ofi, f_ult)
        limite = limite_ofi(razon, via)
        desvio = hs is not None and hs > limite
        results.append({
            'razon': razon, 'nombre': 'FASA' if razon == FASA else 'FSM',
            'ref': r.get('Referencia',''), 'carpeta': r.get('Carpeta',''),
            'via': via,
            'f_ofi': f_ofi, 'f_ult': f_ult,
            'hs': round(hs, 1) if hs else None,
            'limite': limite, 'desvio': desvio,
            'desvio_desc': '', 'parametro': ''
        })
    return results

def procesar_cm_presentados(df):
    results = []
    for _, r in df.iterrows():
        f_tad = parse_date(r.get('TAD SUBIDO'))
        f_ult = parse_date(r.get('Ult evento'))
        hs    = business_hours(f_tad, f_ult)
        desvio = hs is not None and hs > 48
        results.append({
            'carpeta': r.get('CARPETA',''), 'exp': r.get('Expediente',''),
            'f_tad': f_tad, 'f_ult': f_ult,
            'hs': round(hs, 1) if hs else None,
            'desvio': desvio, 'desvio_desc': '', 'parametro': ''
        })
    return results

def procesar_cm_aprobados(df):
    results = []
    for _, r in df.iterrows():
        f1 = parse_date(r.get('Fecha'))
        f2 = parse_date(r.get('Fechadeaprobacion'))
        dias = (f2 - f1).days if f1 and f2 else None
        rango = None
        if dias is not None:
            rango = '0 a 7' if dias <= 7 else ('8 a 15' if dias <= 15 else '+15')
        results.append({
            'carpeta': r.get('CARPETA',''), 'exp': r.get('Expediente',''),
            'f_inicio': f1, 'f_apro': f2,
            'dias': dias, 'rango': rango
        })
    return results

def calcular_kpi(items, con_parametros=False):
    total = len(items)
    if total == 0: return 0, 0, 0
    if con_parametros:
        out = sum(1 for i in items if i['desvio'] and str(i.get('parametro','')).upper() == 'INTERLOG')
    else:
        out = sum(1 for i in items if i['desvio'])
    in_count = total - out
    return round(in_count / total * 100, 2), in_count, out

# ─────────────────────────────────────────────
# CHARTS
# ─────────────────────────────────────────────
def chart_gauge(pct, title):
    color = color_kpi(pct)
    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=pct,
        number={'suffix': '%', 'font': {'size': 36, 'color': color, 'family': 'Barlow Condensed'}},
        title={'text': title, 'font': {'size': 13, 'color': '#9AB0C4', 'family': 'Barlow'}},
        gauge={
            'axis': {'range': [0, 100], 'tickwidth': 1, 'tickcolor': '#6B8099',
                     'tickfont': {'size': 10, 'color': '#6B8099'}},
            'bar': {'color': color, 'thickness': 0.25},
            'bgcolor': '#1A2E48',
            'borderwidth': 0,
            'steps': [
                {'range': [0, 80],  'color': 'rgba(255,61,94,0.15)'},
                {'range': [80, 95], 'color': 'rgba(255,140,66,0.15)'},
                {'range': [95, 100],'color': 'rgba(0,201,167,0.15)'},
            ],
            'threshold': {'line': {'color': '#FFD060', 'width': 3}, 'thickness': 0.8, 'value': 95}
        }
    ))
    fig.update_layout(**CHART_LAYOUT, height=220)
    return fig

def chart_hbar(labels, values, kpi_pcts, colors=None):
    if not labels or not values:
        return go.Figure().update_layout(**CHART_LAYOUT, height=120)
    colors = colors or [COLORS['accent']] * len(labels)
    fig = go.Figure()
    for i, (lbl, val, pct, col) in enumerate(zip(labels, values, kpi_pcts, colors)):
        fig.add_trace(go.Bar(
            y=[lbl], x=[val], orientation='h',
            marker_color=col, marker_line_width=0,
            text=f"  {val}  ({pct:.1f}%)", textposition='outside',
            textfont=dict(color='#F0F4F8', size=11, family='Barlow'),
            name=lbl, showlegend=False
        ))
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(19,34,54,0.8)',
        font=dict(family='Barlow, sans-serif', color='#9AB0C4'),
        margin=dict(l=10, r=10, t=30, b=10),
        height=max(120, len(labels) * 55 + 40),
        barmode='overlay',
    )
    fig.update_xaxes(showticklabels=False, showgrid=False, zeroline=False)
    fig.update_yaxes(tickfont=dict(size=11, family='Barlow', color='#9AB0C4'))
    return fig

def chart_donut(labels, values, colors, title=''):
    fig = go.Figure(go.Pie(
        labels=labels, values=values,
        hole=0.6, marker_colors=colors,
        textinfo='none',
        hovertemplate='%{label}: %{value} (%{percent})<extra></extra>'
    ))
    total = sum(values)
    fig.add_annotation(
        text=f"<b>{total}</b>", x=0.5, y=0.55,
        font=dict(size=28, color='#F0F4F8', family='Barlow Condensed'),
        showarrow=False
    )
    fig.add_annotation(
        text="ops", x=0.5, y=0.35,
        font=dict(size=12, color='#6B8099', family='Barlow'),
        showarrow=False
    )
    fig.update_layout(**CHART_LAYOUT, height=260, showlegend=True,
        legend=dict(orientation='h', yanchor='bottom', y=-0.15,
                    font=dict(size=10, color='#9AB0C4')))
    return fig

def chart_scatter_tiempo(items, limite, title=''):
    hs_vals = [i['hs'] for i in items if i['hs'] is not None]
    if not hs_vals: return None
    colors = [COLORS['red'] if h > limite else COLORS['accent'] for h in hs_vals]
    fig = go.Figure()
    fig.add_hline(y=limite, line_dash='dash', line_color=COLORS['orange'],
                  annotation_text=f'Límite {limite}hs', annotation_position='top right',
                  annotation_font=dict(color=COLORS['orange'], size=11))
    fig.add_trace(go.Scatter(
        x=list(range(len(hs_vals))), y=hs_vals,
        mode='markers',
        marker=dict(color=colors, size=10, line=dict(color='white', width=1.5)),
        hovertemplate='Op %{x}: %{y:.1f} hs<extra></extra>'
    ))
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(19,34,54,0.8)',
        font=dict(family='Barlow, sans-serif', color='#9AB0C4'),
        margin=dict(l=10, r=10, t=30, b=10),
        height=220,
    )
    fig.update_xaxes(showticklabels=False, title='Operaciones')
    fig.update_yaxes(title='Horas hábiles')
    return fig

def chart_stacked_canales(nombre, items):
    vias = ['AVION', 'MARITIMO', 'CAMION']
    verde_v, naranja_v, rojo_v, labels = [], [], [], []
    for via in vias:
        via_items = [i for i in items if i['via'] == via]
        if not via_items: continue
        by_canal = Counter(i['canal'] for i in via_items)
        verde_v.append(by_canal.get('VERDE', 0))
        naranja_v.append(by_canal.get('NARANJA', 0))
        rojo_v.append(by_canal.get('ROJO', 0))
        labels.append(f"{via} (n={len(via_items)})")

    if not labels: return None
    fig = go.Figure()
    fig.add_trace(go.Bar(name='Verde',   x=labels, y=verde_v,   marker_color=COLORS['verde'],   text=verde_v,   textposition='inside', textfont=dict(color='white', size=11)))
    fig.add_trace(go.Bar(name='Naranja', x=labels, y=naranja_v, marker_color=COLORS['naranja'], text=naranja_v, textposition='inside', textfont=dict(color='white', size=11)))
    fig.add_trace(go.Bar(name='Rojo',    x=labels, y=rojo_v,    marker_color=COLORS['rojo'],    text=rojo_v,    textposition='inside', textfont=dict(color='white', size=11)))
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(19,34,54,0.8)',
        font=dict(family='Barlow, sans-serif', color='#9AB0C4'),
        margin=dict(l=10, r=10, t=30, b=10),
        barmode='stack', height=280,
        legend=dict(orientation='h', yanchor='bottom', y=1.02,
                    font=dict(color='#9AB0C4', size=11)),
    )
    fig.update_xaxes(tickfont=dict(size=11, color='#9AB0C4'))
    return fig

# ─────────────────────────────────────────────
# GENERAR EXCEL DE DESVÍOS ESTILIZADO
# ─────────────────────────────────────────────
def generar_excel_desvios(lib_items, ofi_items, cm_pre_items, mes='MES'):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DARK_BG = "0D1B2A"; MID_BG = "1B2B3E"; CARD = "132236"
    ACCENT  = "008B74"; GOLD = "FFD060";  ROJO = "FF3D5E"
    WHITE   = "FFFFFF"; LGRAY = "9AB0C4"; PURPLE = "5B21B6"

    def hfill(c): return PatternFill("solid", fgColor=c)
    def bw(s=10):  return Font(bold=True, color=WHITE, size=s, name="Calibri")
    def nw(s=10):  return Font(color=WHITE, size=s, name="Calibri")
    def nr(s=10):  return Font(color=ROJO, size=s, name="Calibri", bold=True)
    def ng(s=10):  return Font(color=GOLD, size=s, name="Calibri", bold=True)
    def cen():     return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lft():     return Alignment(horizontal="left",   vertical="center", wrap_text=True)
    def brd():
        s = Side(border_style="thin", color="1B2B3E")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = Workbook()

    def make_sheet(ws, title_text, headers, rows_data, header_color, col_widths,
                   edit_cols, ref_col_idx=None):
        # Título
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        ws["A1"] = title_text
        ws["A1"].font = bw(13); ws["A1"].fill = hfill(DARK_BG)
        ws["A1"].alignment = cen(); ws.row_dimensions[1].height = 30

        # Subtítulo
        ws.merge_cells(f"A2:{get_column_letter(len(headers))}2")
        ws["A2"] = "⚠️  Completar columnas DESVÍO y PARÁMETRO — Si Parámetro ≠ INTERLOG la operación se considera IN"
        ws["A2"].font = Font(color=GOLD, size=9, name="Calibri")
        ws["A2"].fill = hfill(MID_BG); ws["A2"].alignment = cen()
        ws.row_dimensions[2].height = 20

        # Headers
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(3, ci, h)
            cell.font = bw(10); cell.fill = hfill(header_color)
            cell.alignment = cen(); cell.border = brd()
        ws.row_dimensions[3].height = 25

        # Datos
        if not rows_data:
            ws.merge_cells(f"A4:{get_column_letter(len(headers))}4")
            ws["A4"] = "✅  Sin desvíos detectados — KPI 100%"
            ws["A4"].font = Font(color="00C9A7", size=11, bold=True, name="Calibri")
            ws["A4"].fill = hfill(MID_BG); ws["A4"].alignment = cen()
            ws.row_dimensions[4].height = 30
        else:
            for ri, row_vals in enumerate(rows_data, 4):
                fill_c = CARD if ri % 2 == 0 else MID_BG
                for ci, val in enumerate(row_vals, 1):
                    cell = ws.cell(ri, ci, val)
                    cell.border = brd()
                    cell.alignment = cen() if ci > 1 else lft()
                    if ci in edit_cols:
                        cell.fill = hfill("1A3A2A"); cell.font = ng(10)
                    elif ci == ref_col_idx:
                        cell.fill = hfill(fill_c); cell.font = bw(10)
                    else:
                        cell.fill = hfill(fill_c)
                        # Hs en rojo
                        if headers[ci-1] in ['Días Hábiles']:
                            cell.font = nr(10)
                        else:
                            cell.font = nw(10)
                ws.row_dimensions[ri].height = 20

        # Anchos
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A4"

    # ── HOJA 1: LIBERADAS ──
    ws1 = wb.active; ws1.title = "LIBERADAS - DESVÍOS"
    lib_desvios = [i for i in lib_items if i['desvio']]
    rows_lib = [[
        'FASA' if i['razon'] == FASA else 'FSM',
        i['ref'], str(i['carpeta']), i['via'], i['canal'],
        i['f_ofi'].strftime('%d/%m/%Y') if i['f_ofi'] else '',
        i['f_cancel'].strftime('%d/%m/%Y %H:%M') if i['f_cancel'] else '',
        i['hs'], i['limite'], '', ''
    ] for i in lib_desvios]
    make_sheet(ws1,
        f"LIBERADAS {mes} — OPERACIONES CON DESVÍO",
        ["Razón Social","Referencia","Carpeta","Vía","Canal",
         "F. Oficialización","F. Cancelada","Días Hábiles","Límite (días)","DESVÍO ✏️","PARÁMETRO ✏️"],
        rows_lib, ACCENT, [28,16,12,12,10,18,22,18,12,35,25],
        edit_cols=[10,11], ref_col_idx=2
    )

    # ── HOJA 2: OFICIALIZADOS ──
    ws2 = wb.create_sheet("OFICIALIZADOS - DESVÍOS")
    ofi_desvios = [i for i in ofi_items if i['desvio']]
    rows_ofi = [[
        'FASA' if i['razon'] == FASA else 'FSM',
        i['ref'], str(i['carpeta']), i['via'],
        i['f_ofi'].strftime('%d/%m/%Y') if i['f_ofi'] else '',
        i['f_ult'].strftime('%d/%m/%Y %H:%M') if i['f_ult'] else '',
        i['hs'], i['limite'], '', ''
    ] for i in ofi_desvios]
    make_sheet(ws2,
        f"OFICIALIZADOS {mes} — OPERACIONES CON DESVÍO",
        ["Razón Social","Referencia","Carpeta","Vía",
         "F. Oficialización","Último Evento","Hs Transcurridas","Límite (hs)","DESVÍO ✏️","PARÁMETRO ✏️"],
        rows_ofi, "005F52", [28,16,12,12,18,22,18,12,35,25],
        edit_cols=[9,10], ref_col_idx=2
    )

    # ── HOJA 3: CM PRESENTADOS ──
    ws3 = wb.create_sheet("CM PRESENTADOS - DESVÍOS")
    cm_desvios = [i for i in cm_pre_items if i['desvio']]
    rows_cm = [[
        str(i['carpeta']), str(i['exp']),
        i['f_tad'].strftime('%d/%m/%Y') if i['f_tad'] else '',
        i['f_ult'].strftime('%d/%m/%Y') if i['f_ult'] else '',
        i['hs'], 48, '', ''
    ] for i in cm_desvios]
    make_sheet(ws3,
        f"CM PRESENTADOS {mes} — OPERACIONES CON DESVÍO",
        ["Carpeta","Expediente","TAD Subido","Último Evento",
         "Hs Transcurridas","Límite (hs)","DESVÍO ✏️","PARÁMETRO ✏️"],
        rows_cm, PURPLE, [12,28,18,18,18,12,35,25],
        edit_cols=[7,8], ref_col_idx=2
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
# EXPORT DASHBOARD COMPLETO A EXCEL
# ─────────────────────────────────────────────
def export_dashboard_excel(lib_items, ofi_items, cm_pre_items, cm_apr_items, mes='MES'):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DARK='0D1B2A'; MID='1B2B3E'; CARD='132236'; CARD2='1A2E48'
    ACCENT='008B74'; GOLD='FFD060'; ROJO_C='FF3D5E'; WHITE='FFFFFF'
    ORANGE='CC6600'; VERDE_C='007A65'; LGRAY='9AB0C4'

    def hf(c): return PatternFill("solid", fgColor=c)
    def fw(s=10, bold=False): return Font(bold=bold, color=WHITE, size=s, name="Calibri")
    def fg(s=10): return Font(bold=True, color=GOLD, size=s, name="Calibri")
    def fr(s=10): return Font(bold=True, color=ROJO_C, size=s, name="Calibri")
    def fa(s=10): return Font(bold=True, color='00C9A7', size=s, name="Calibri")
    def cen(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lft(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
    def brd():
        s = Side(border_style="thin", color="1B2B3E")
        return Border(left=s, right=s, top=s, bottom=s)

    def write_title(ws, text, ncols, row=1, height=30):
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws.cell(row, 1, text)
        c.font = fw(13, True); c.fill = hf(DARK); c.alignment = cen()
        ws.row_dimensions[row].height = height

    def write_headers(ws, headers, row, fill_hex):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row, ci, h)
            c.font = fw(10, True); c.fill = hf(fill_hex)
            c.alignment = cen(); c.border = brd()
        ws.row_dimensions[row].height = 24

    def write_row(ws, row_num, values, row_fill):
        for ci, val in enumerate(values, 1):
            c = ws.cell(row_num, ci, val)
            c.border = brd(); c.alignment = cen()
            c.fill = hf(row_fill); c.font = fw(10)

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def kpi_font(pct):
        if pct >= 95: return fa()
        if pct >= 80: return Font(bold=True, color=ORANGE, size=10, name="Calibri")
        return fr()

    def param_font(param):
        if str(param).upper() == 'INTERLOG': return fr()
        if param: return fa()
        return fw(10)

    wb = Workbook()

    # ══════════════════════════════════════════
    # HOJA 1 — RESUMEN EJECUTIVO
    # ══════════════════════════════════════════
    ws1 = wb.active; ws1.title = "RESUMEN EJECUTIVO"
    ws1.sheet_view.showGridLines = False
    write_title(ws1, f"RESUMEN EJECUTIVO · KPI {mes}", 7)
    write_headers(ws1, ["Proceso","Razón Social","Total Ops","IN","OUT","KPI %","Target"], 2, ACCENT)

    filas_res = []
    for proc, items, nombre in [
        ("LIBERACIÓN", lib_items, "FASA"), ("LIBERACIÓN", lib_items, "FSM"),
        ("OFICIALIZACIÓN", ofi_items, "FASA"), ("OFICIALIZACIÓN", ofi_items, "FSM"),
    ]:
        sub = [i for i in items if i['nombre'] == nombre]
        pct, in_c, out_c = calcular_kpi(sub, True)
        filas_res.append([proc, nombre, len(sub), in_c, out_c, f"{pct:.1f}%", "95%"])
    sub_cm = cm_pre_items
    pct_cm, in_cm, out_cm = calcular_kpi(sub_cm, True)
    filas_res.append(["CM PRESENTADOS", "FASA+FSM", len(sub_cm), in_cm, out_cm, f"{pct_cm:.1f}%", "95%"])

    for ri, vals in enumerate(filas_res, 3):
        fill_c = CARD if ri % 2 == 1 else MID
        write_row(ws1, ri, vals, fill_c)
        pct_val = float(vals[5].replace('%',''))
        ws1.cell(ri, 6).font = kpi_font(pct_val)
    set_col_widths(ws1, [20, 14, 12, 8, 8, 10, 10])
    ws1.freeze_panes = "A3"

    # ══════════════════════════════════════════
    # HOJA 2 — LIBERACIONES DETALLE
    # ══════════════════════════════════════════
    ws2 = wb.create_sheet("LIBERACIONES - DETALLE")
    ws2.sheet_view.showGridLines = False
    write_title(ws2, f"LIBERACIONES · DETALLE POR SOCIEDAD, VÍA Y CANAL · {mes}", 12)
    headers_lib = ["Razón Social","Vía","Canal","Total Ops","IN","OUT","KPI %",
                   "Prom Días","Límite Días","Referencia","DESVÍO","PARÁMETRO"]
    write_headers(ws2, headers_lib, 2, ACCENT)

    ri = 3
    vias_orden = ['AVION','CAMION','MARITIMO']
    canales_orden = ['VERDE','NARANJA','ROJO']
    for nombre in ['FASA','FSM']:
        for via in vias_orden:
            for canal in canales_orden:
                ops = [i for i in lib_items if i['nombre']==nombre and i['via']==via and i['canal']==canal]
                if not ops: continue
                pct, in_c, out_c = calcular_kpi(ops, True)
                dias_vals = [i['hs'] for i in ops if i['hs'] is not None]
                avg_dias = round(np.mean(dias_vals), 1) if dias_vals else ''
                limite = ops[0]['limite']
                fill_c = CARD if ri % 2 == 1 else MID

                # Fila resumen del grupo
                ws2.cell(ri, 1, nombre).font = fw(10, True)
                ws2.cell(ri, 1).fill = hf(fill_c); ws2.cell(ri, 1).border = brd(); ws2.cell(ri, 1).alignment = cen()
                for ci, val in enumerate([via, canal, len(ops), in_c, out_c, f"{pct:.1f}%", avg_dias, limite, '', '', ''], 2):
                    c = ws2.cell(ri, ci, val)
                    c.fill = hf(fill_c); c.border = brd(); c.alignment = cen()
                    if ci == 7: c.font = kpi_font(pct)
                    elif ci == 3:
                        cc = {'VERDE': VERDE_C, 'NARANJA': ORANGE, 'ROJO': '880022'}.get(canal, MID)
                        c.fill = hf(cc); c.font = fw(10, True)
                    else: c.font = fw(10)
                ws2.row_dimensions[ri].height = 20
                ri += 1

                # Filas detalle — solo desvíos
                for op in [o for o in ops if o['desvio']]:
                    fill_d = '1A0A0A'
                    for ci, val in enumerate(['', '', '', '', '', '', '', '', '',
                                              op['ref'], op.get('desvio_desc',''), op.get('parametro','')], 1):
                        c = ws2.cell(ri, ci, val)
                        c.fill = hf(fill_d); c.border = brd(); c.alignment = lft() if ci >= 10 else cen()
                        if ci == 12: c.font = param_font(op.get('parametro',''))
                        elif ci == 11: c.font = fg(9)
                        else: c.font = fw(9)
                    ws2.row_dimensions[ri].height = 18
                    ri += 1

    set_col_widths(ws2, [14,12,10,10,8,8,10,10,10,18,35,20])
    ws2.freeze_panes = "A3"

    # ══════════════════════════════════════════
    # HOJA 3 — OFICIALIZACIONES DETALLE
    # ══════════════════════════════════════════
    ws3 = wb.create_sheet("OFICIALIZACIONES - DETALLE")
    ws3.sheet_view.showGridLines = False
    write_title(ws3, f"OFICIALIZACIONES · DETALLE POR SOCIEDAD Y VÍA · {mes}", 10)
    headers_ofi = ["Razón Social","Vía","Total Ops","IN","OUT","KPI %",
                   "Prom Hs","Límite Hs","DESVÍO","PARÁMETRO"]
    write_headers(ws3, headers_ofi, 2, '005F52')

    ri = 3
    for nombre in ['FASA','FSM']:
        for via in vias_orden:
            ops = [i for i in ofi_items if i['nombre']==nombre and i['via']==via]
            if not ops: continue
            pct, in_c, out_c = calcular_kpi(ops, True)
            hs_vals = [i['hs'] for i in ops if i['hs'] is not None]
            avg_hs = round(np.mean(hs_vals), 1) if hs_vals else ''
            limite = ops[0]['limite']
            fill_c = CARD if ri % 2 == 1 else MID

            ws3.cell(ri, 1, nombre).font = fw(10, True)
            ws3.cell(ri, 1).fill = hf(fill_c); ws3.cell(ri, 1).border = brd(); ws3.cell(ri, 1).alignment = cen()
            for ci, val in enumerate([via, len(ops), in_c, out_c, f"{pct:.1f}%", avg_hs, limite, '', ''], 2):
                c = ws3.cell(ri, ci, val)
                c.fill = hf(fill_c); c.border = brd(); c.alignment = cen()
                if ci == 6: c.font = kpi_font(pct)
                else: c.font = fw(10)
            ws3.row_dimensions[ri].height = 20
            ri += 1

            for op in [o for o in ops if o['desvio']]:
                fill_d = '1A0A0A'
                for ci, val in enumerate(['','','','','','','','',
                                          op.get('desvio_desc',''), op.get('parametro','')], 1):
                    c = ws3.cell(ri, ci, val)
                    c.fill = hf(fill_d); c.border = brd(); c.alignment = lft() if ci >= 9 else cen()
                    if ci == 10: c.font = param_font(op.get('parametro',''))
                    elif ci == 9: c.font = fg(9)
                    else: c.font = fw(9)
                ws3.row_dimensions[ri].height = 18
                ri += 1

    set_col_widths(ws3, [14,12,10,8,8,10,10,10,35,20])
    ws3.freeze_panes = "A3"

    # ══════════════════════════════════════════
    # HOJA 4 — CM PRESENTADOS
    # ══════════════════════════════════════════
    ws4 = wb.create_sheet("CM PRESENTADOS")
    ws4.sheet_view.showGridLines = False
    write_title(ws4, f"CERTIFICADOS MINEROS PRESENTADOS · {mes}", 8)
    headers_cm = ["Carpeta","Expediente","TAD Subido","Último Evento",
                  "Hs Transcurridas","Límite (hs)","DESVÍO","PARÁMETRO"]
    write_headers(ws4, headers_cm, 2, '5B21B6')

    for ri, i in enumerate(cm_pre_items, 3):
        fill_c = CARD if ri % 2 == 1 else MID
        vals = [
            str(i.get('carpeta','')), str(i['exp']),
            i['f_tad'].strftime('%d/%m/%Y') if i.get('f_tad') else '',
            i['f_ult'].strftime('%d/%m/%Y') if i.get('f_ult') else '',
            i['hs'], 48,
            i.get('desvio_desc','') if i['desvio'] else '',
            i.get('parametro','') if i['desvio'] else '',
        ]
        for ci, val in enumerate(vals, 1):
            c = ws4.cell(ri, ci, val)
            c.fill = hf('1A0A0A' if i['desvio'] else fill_c)
            c.border = brd(); c.alignment = lft() if ci >= 7 else cen()
            if ci == 8: c.font = param_font(i.get('parametro',''))
            elif ci == 7 and i['desvio']: c.font = fg(10)
            else: c.font = fw(10)
        ws4.row_dimensions[ri].height = 20

    set_col_widths(ws4, [12,28,16,16,16,12,35,20])
    ws4.freeze_panes = "A3"

    # ══════════════════════════════════════════
    # HOJA 5 — CM APROBADOS
    # ══════════════════════════════════════════
    if cm_apr_items:
        ws5 = wb.create_sheet("CM APROBADOS")
        ws5.sheet_view.showGridLines = False
        write_title(ws5, f"CERTIFICADOS MINEROS APROBADOS · {mes}", 4)
        write_headers(ws5, ["Carpeta","Expediente","Días para aprobación","Rango"], 2, '5B21B6')
        rango_colors = {'0 a 7': VERDE_C, '8 a 15': ORANGE, '+15': '880022'}
        for ri, i in enumerate(cm_apr_items, 3):
            fill_c = CARD if ri % 2 == 1 else MID
            vals = [str(i.get('carpeta','')), str(i['exp']),
                    i['dias'] if i['dias'] is not None else '', i['rango'] or '']
            for ci, val in enumerate(vals, 1):
                c = ws5.cell(ri, ci, val)
                c.fill = hf(fill_c); c.border = brd(); c.alignment = cen()
                if ci == 4 and i['rango']:
                    c.fill = hf(rango_colors.get(i['rango'], MID))
                    c.font = fw(10, True)
                else:
                    c.font = fw(10)
            ws5.row_dimensions[ri].height = 20
        set_col_widths(ws5, [12,28,22,14])
        ws5.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────
# SECCIÓN KPI
# ─────────────────────────────────────────────
def render_via_section(nombre, via, lib_items, ofi_items, con_param, key_prefix):
    """Renderiza una sección de Oficialización + Liberación para una sociedad + vía"""
    lib_r = [i for i in lib_items if i['nombre'] == nombre and i['via'] == via]
    ofi_r = [i for i in ofi_items if i['nombre'] == nombre and i['via'] == via]

    if not lib_r and not ofi_r:
        st.markdown(f'<div class="alert-info">Sin operaciones de vía {via} para {nombre}.</div>', unsafe_allow_html=True)
        return

    via_emoji = {'AVION': '✈️', 'CAMION': '🚛', 'MARITIMO': '🚢'}.get(via, '📦')
    st.markdown(f"""
    <div style="background:#132236; border-left:4px solid #FFD060;
         padding:0.5rem 1rem; margin:0.8rem 0 0.4rem; border-radius:0 6px 6px 0;">
        <span style="font-family:'Barlow Condensed',sans-serif; font-size:1.1rem;
        font-weight:800; color:#FFD060; letter-spacing:2px;">{via_emoji} {nombre} · {via}</span>
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    # ── OFICIALIZACIÓN ──
    with col1:
        st.markdown('<div style="color:#9AB0C4; font-size:0.8rem; font-weight:600; letter-spacing:1px; text-transform:uppercase; margin-bottom:0.3rem;">OFICIALIZACIÓN</div>', unsafe_allow_html=True)
        if ofi_r:
            pct_ofi, in_ofi, out_ofi = calcular_kpi(ofi_r, con_param)
            c1, c2, c3 = st.columns(3)
            c1.plotly_chart(chart_gauge(pct_ofi, "KPI IN"), use_container_width=True, key=f"{key_prefix}_gauge_ofi")
            params = Counter(i['parametro'] or 'S/DESVIO' for i in ofi_r if not i['desvio'])
            params.update({i['parametro'] or 'PENDIENTE': 1 for i in ofi_r if i['desvio']})
            lbls = list(params.keys()); vals = list(params.values())
            cols_bar = [COLORS['accent'] if l=='S/DESVIO' else (COLORS['red'] if l=='INTERLOG' else COLORS['orange']) for l in lbls]
            kpi_p = [v/len(ofi_r)*100 for v in vals]
            with c2:
                st.plotly_chart(chart_hbar(lbls, vals, kpi_p, cols_bar), use_container_width=True, key=f"{key_prefix}_hbar_ofi")
            with c3:
                st.markdown(metric_html(str(len(ofi_r)), "Total", None, 'accent'), unsafe_allow_html=True)
                st.markdown(metric_html(str(in_ofi), "IN", None, 'accent'), unsafe_allow_html=True)
                st.markdown(metric_html(str(out_ofi), "OUT", None, 'red' if out_ofi else 'accent'), unsafe_allow_html=True)
        else:
            st.markdown('<div class="alert-info">Sin oficializaciones para esta vía.</div>', unsafe_allow_html=True)

    # ── LIBERACIÓN ──
    with col2:
        st.markdown('<div style="color:#9AB0C4; font-size:0.8rem; font-weight:600; letter-spacing:1px; text-transform:uppercase; margin-bottom:0.3rem;">LIBERACIÓN</div>', unsafe_allow_html=True)
        if lib_r:
            pct_lib, in_lib, out_lib = calcular_kpi(lib_r, con_param)
            c1, c2, c3 = st.columns(3)
            c1.plotly_chart(chart_gauge(pct_lib, "KPI IN"), use_container_width=True, key=f"{key_prefix}_gauge_lib")
            params = Counter(i['parametro'] or 'S/DESVIO' for i in lib_r if not i['desvio'])
            params.update({i['parametro'] or 'PENDIENTE': 1 for i in lib_r if i['desvio']})
            lbls = list(params.keys()); vals = list(params.values())
            cols_bar = [COLORS['accent'] if l=='S/DESVIO' else (COLORS['red'] if l=='INTERLOG' else COLORS['orange']) for l in lbls]
            kpi_p = [v/len(lib_r)*100 for v in vals]
            with c2:
                st.plotly_chart(chart_hbar(lbls, vals, kpi_p, cols_bar), use_container_width=True, key=f"{key_prefix}_hbar_lib")
            with c3:
                st.markdown(metric_html(str(len(lib_r)), "Total", None, 'accent'), unsafe_allow_html=True)
                st.markdown(metric_html(str(in_lib), "IN", None, 'accent'), unsafe_allow_html=True)
                st.markdown(metric_html(str(out_lib), "OUT", None, 'red' if out_lib else 'accent'), unsafe_allow_html=True)

            # Canal Verde Avión
            if via == 'AVION':
                va = [i for i in lib_r if i['canal'] == 'VERDE']
                if va:
                    pct_va, in_va, out_va = calcular_kpi(va, con_param)
                    st.markdown(f'<div class="section-header" style="border-color:#00C9A7; margin-top:0.8rem;">CANAL VERDE · {nombre}</div>', unsafe_allow_html=True)
                    ca, cb, cc, cd = st.columns(4)
                    ca.markdown(metric_html(f"{pct_va:.1f}%", "KPI IN", "Target: 95%", 'accent' if pct_va>=95 else 'orange'), unsafe_allow_html=True)
                    cb.markdown(metric_html(str(len(va)), "Total ops", "Canal Verde", 'accent'), unsafe_allow_html=True)
                    cc.markdown(metric_html(str(in_va), "IN", None, 'accent'), unsafe_allow_html=True)
                    cd.markdown(metric_html(str(out_va), "OUT", None, 'red' if out_va else 'accent'), unsafe_allow_html=True)
        else:
            st.markdown('<div class="alert-info">Sin liberaciones para esta vía.</div>', unsafe_allow_html=True)

    st.markdown('<hr style="border-color:rgba(107,128,153,0.2); margin:1rem 0;">', unsafe_allow_html=True)


def render_kpi_section(nombre, lib_items, ofi_items, con_param=False):
    n = nombre.lower()
    vias_presentes = sorted(set(
        [i['via'] for i in lib_items if i['nombre'] == nombre] +
        [i['via'] for i in ofi_items if i['nombre'] == nombre]
    ), key=lambda x: ['AVION','CAMION','MARITIMO'].index(x) if x in ['AVION','CAMION','MARITIMO'] else 99)

    for via in vias_presentes:
        render_via_section(nombre, via, lib_items, ofi_items, con_param, key_prefix=f"{n}_{via.lower()}")

# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────

# Header
st.markdown("""
<div style="background: linear-gradient(135deg, #0F2040 0%, #132236 100%);
     border-bottom: 1px solid rgba(0,201,167,0.2);
     padding: 1.5rem 2rem; margin: -1rem -1rem 1.5rem; border-radius: 0 0 12px 12px;">
  <div style="display:flex; align-items:center; gap:1rem;">
    <div>
      <div style="font-family:'Barlow Condensed',sans-serif; font-size:2rem;
           font-weight:800; color:#00C9A7; letter-spacing:3px;">INTERLOG</div>
      <div style="font-family:'Barlow Condensed',sans-serif; font-size:1rem;
           color:#6B8099; letter-spacing:2px; text-transform:uppercase;">KPI Dashboard · Comercio Exterior</div>
    </div>
    <div style="margin-left:auto; text-align:right;">
      <div style="font-size:0.75rem; color:#6B8099; text-transform:uppercase; letter-spacing:1px;">Reporte Mensual</div>
      <div style="font-family:'Barlow Condensed',sans-serif; font-size:1.2rem;
           color:#F0F4F8; font-weight:600;">FASA / FSM</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── ESTADO DE SESIÓN ──
if 'step' not in st.session_state:
    st.session_state.step = 1
if 'lib_items' not in st.session_state:
    st.session_state.lib_items = []
if 'ofi_items' not in st.session_state:
    st.session_state.ofi_items = []
if 'cm_pre_items' not in st.session_state:
    st.session_state.cm_pre_items = []
if 'cm_apr_items' not in st.session_state:
    st.session_state.cm_apr_items = []
if 'mes' not in st.session_state:
    st.session_state.mes = ''

# ── STEP INDICATOR ──
steps = ["📁 Cargar archivos", "⚠️ Revisar desvíos", "📊 Dashboard", "📥 Exportar"]
cols_steps = st.columns(4)
for i, (col, label) in enumerate(zip(cols_steps, steps)):
    active = st.session_state.step == i+1
    done   = st.session_state.step > i+1
    bg = '#00C9A7' if active else ('#007A65' if done else '#132236')
    tc = '#0A1628' if active else ('#F0F4F8' if done else '#6B8099')
    col.markdown(f"""
    <div style="background:{bg}; color:{tc}; border-radius:6px; padding:0.5rem;
         text-align:center; font-family:'Barlow Condensed',sans-serif;
         font-size:0.85rem; font-weight:700; letter-spacing:0.5px;">
        {label}
    </div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ══════════════════════════════════════════════
# STEP 1 — CARGAR ARCHIVOS
# ══════════════════════════════════════════════
if st.session_state.step == 1:
    st.markdown('<div class="section-header">PASO 1 · CARGAR ARCHIVOS EXCEL</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="alert-info">
    Subí los 4 archivos Excel del mes. El sistema procesará automáticamente todas las métricas.
    </div><br>
    """, unsafe_allow_html=True)

    # Estilos para los uploaders — mayor contraste
    st.markdown("""
    <style>
    [data-testid="stFileUploader"] {
        background: #1A2E48 !important;
        border: 2px solid #00C9A7 !important;
        border-radius: 8px !important;
        padding: 0.5rem !important;
    }
    [data-testid="stFileUploader"] label {
        color: #F0F4F8 !important;
        font-weight: 700 !important;
        font-size: 1rem !important;
        letter-spacing: 0.5px !important;
    }
    [data-testid="stFileUploaderDropzone"] {
        background: #0F2040 !important;
        border: 2px dashed #00C9A7 !important;
        color: #F0F4F8 !important;
    }
    [data-testid="stFileUploaderDropzone"] p {
        color: #9AB0C4 !important;
        font-weight: 500 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    mes = st.text_input("📅 Mes del reporte (ej: DICIEMBRE 2025)", value=st.session_state.mes)

    col1, col2 = st.columns(2)
    with col1:
        f_lib = st.file_uploader("📦 LIBERADAS", type=['xlsx'], key='lib')
        f_ofi = st.file_uploader("📋 OFICIALIZADOS", type=['xlsx'], key='ofi')
    with col2:
        f_cm_pre = st.file_uploader("📜 CM PRESENTADOS", type=['xlsx'], key='cmpre')
        f_cm_apr = st.file_uploader("✅ CM APROBADOS", type=['xlsx'], key='cmapr')

    archivos_ok = all([f_lib, f_ofi, f_cm_pre, f_cm_apr])

    if archivos_ok:
        st.markdown('<div class="alert-success">✅ Los 4 archivos cargados correctamente</div><br>', unsafe_allow_html=True)

    if st.button("▶  PROCESAR Y CONTINUAR", disabled=not archivos_ok):
        with st.spinner("Procesando datos..."):
            df_lib     = pd.read_excel(f_lib)
            df_ofi     = pd.read_excel(f_ofi)
            df_cm_pre  = pd.read_excel(f_cm_pre)
            df_cm_apr  = pd.read_excel(f_cm_apr)

            st.session_state.lib_items     = procesar_liberadas(df_lib)
            st.session_state.ofi_items     = procesar_oficializados(df_ofi)
            st.session_state.cm_pre_items  = procesar_cm_presentados(df_cm_pre)
            st.session_state.cm_apr_items  = procesar_cm_aprobados(df_cm_apr)
            st.session_state.mes           = mes
            st.session_state.step          = 2
            st.rerun()

# ══════════════════════════════════════════════
# STEP 2 — REVISAR DESVÍOS
# ══════════════════════════════════════════════
elif st.session_state.step == 2:
    lib_items    = st.session_state.lib_items
    ofi_items    = st.session_state.ofi_items
    cm_pre_items = st.session_state.cm_pre_items

    desvios_lib = [i for i in lib_items if i['desvio']]
    desvios_ofi = [i for i in ofi_items if i['desvio']]
    desvios_cm  = [i for i in cm_pre_items if i['desvio']]
    total_desvios = len(desvios_lib) + len(desvios_ofi) + len(desvios_cm)

    st.markdown('<div class="section-header">PASO 2 · INSTANCIA DE REVISIÓN DE DESVÍOS</div>', unsafe_allow_html=True)

    if total_desvios == 0:
        st.markdown('<div class="alert-success">✅ Sin desvíos detectados en ningún proceso. KPI 100% en todo.</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("▶  IR AL DASHBOARD"):
            st.session_state.step = 3
            st.rerun()
    else:
        # ── SUB-ESTADO: ¿ya descargó o está subiendo? ──
        if 'desvio_sub_step' not in st.session_state:
            st.session_state.desvio_sub_step = 'descargar'

        # ──────────────────────────────────────────
        # SUB-STEP A: DESCARGAR EXCEL
        # ──────────────────────────────────────────
        if st.session_state.desvio_sub_step == 'descargar':
            st.markdown(f"""
            <div class="alert-warn">
            ⚠️  Se detectaron <b>{total_desvios} operaciones fuera del rango</b>.
            Descargá el Excel, completá las columnas <b>DESVÍO</b> y <b>PARÁMETRO</b>, y volvé a subirlo.<br><br>
            <small>💡 Si el Parámetro <b>no es INTERLOG</b> → la operación se considera <b>IN</b> automáticamente.</small>
            </div><br>
            """, unsafe_allow_html=True)

            # Generar Excel estilizado
            excel_buf = generar_excel_desvios(lib_items, ofi_items, cm_pre_items,
                                               st.session_state.mes or 'MES')

            c1, c2, c3 = st.columns([2, 2, 3])
            with c1:
                st.download_button(
                    label="⬇  DESCARGAR EXCEL DE DESVÍOS",
                    data=excel_buf,
                    file_name=f"DESVIOS_{(st.session_state.mes or 'MES').replace(' ','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            with c2:
                if st.button("✅  YA LO COMPLETÉ — SUBIR AHORA", use_container_width=True):
                    st.session_state.desvio_sub_step = 'subir'
                    st.rerun()
            with c3:
                if st.button("◀  VOLVER A CARGAR ARCHIVOS", use_container_width=True):
                    st.session_state.step = 1
                    st.session_state.desvio_sub_step = 'descargar'
                    st.rerun()

        # ──────────────────────────────────────────
        # SUB-STEP B: SUBIR EXCEL COMPLETADO
        # ──────────────────────────────────────────
        elif st.session_state.desvio_sub_step == 'subir':
            st.markdown("""
            <div class="alert-info">
            📂 Subí el Excel de desvíos que ya completaste. El sistema leerá los campos
            <b>DESVÍO</b> y <b>PARÁMETRO</b> de cada hoja y recalculará los KPIs.
            </div><br>
            """, unsafe_allow_html=True)

            f_desvios = st.file_uploader(
                "📋 Excel de Desvíos completado",
                type=['xlsx'], key='desvios_upload'
            )

            if f_desvios:
                try:
                    wb_dev = pd.ExcelFile(f_desvios)
                    errores = []

                    # ── Leer LIBERADAS ──
                    if 'LIBERADAS - DESVÍOS' in wb_dev.sheet_names:
                        df_dev_lib = wb_dev.parse('LIBERADAS - DESVÍOS', skiprows=2)
                        # Normalizar columnas
                        df_dev_lib.columns = [str(c).strip() for c in df_dev_lib.columns]
                        col_desv  = next((c for c in df_dev_lib.columns if 'DESVÍO' in c or 'DESVIO' in c.upper()), None)
                        col_param = next((c for c in df_dev_lib.columns if 'PARÁMETRO' in c or 'PARAMETRO' in c.upper()), None)
                        col_ref   = next((c for c in df_dev_lib.columns if 'REFERENCIA' in c.upper() or 'REF' in c.upper()), None)

                        if col_ref and col_param:
                            ref_map = {}
                            for _, row in df_dev_lib.iterrows():
                                ref = str(row.get(col_ref, '')).strip()
                                if ref:
                                    ref_map[ref] = {
                                        'desvio_desc': str(row.get(col_desv, '') or '').strip(),
                                        'parametro':   str(row.get(col_param,'') or '').strip()
                                    }
                            for item in lib_items:
                                if item['ref'] in ref_map:
                                    item['desvio_desc'] = ref_map[item['ref']]['desvio_desc']
                                    item['parametro']   = ref_map[item['ref']]['parametro']

                    # ── Leer OFICIALIZADOS ──
                    if 'OFICIALIZADOS - DESVÍOS' in wb_dev.sheet_names:
                        df_dev_ofi = wb_dev.parse('OFICIALIZADOS - DESVÍOS', skiprows=2)
                        df_dev_ofi.columns = [str(c).strip() for c in df_dev_ofi.columns]
                        col_desv  = next((c for c in df_dev_ofi.columns if 'DESVÍO' in c or 'DESVIO' in c.upper()), None)
                        col_param = next((c for c in df_dev_ofi.columns if 'PARÁMETRO' in c or 'PARAMETRO' in c.upper()), None)
                        col_ref   = next((c for c in df_dev_ofi.columns if 'REFERENCIA' in c.upper() or 'REF' in c.upper()), None)

                        if col_ref and col_param:
                            ref_map = {}
                            for _, row in df_dev_ofi.iterrows():
                                ref = str(row.get(col_ref, '')).strip()
                                if ref:
                                    ref_map[ref] = {
                                        'desvio_desc': str(row.get(col_desv, '') or '').strip(),
                                        'parametro':   str(row.get(col_param,'') or '').strip()
                                    }
                            for item in ofi_items:
                                if item['ref'] in ref_map:
                                    item['desvio_desc'] = ref_map[item['ref']]['desvio_desc']
                                    item['parametro']   = ref_map[item['ref']]['parametro']

                    # ── Leer CM PRESENTADOS ──
                    if 'CM PRESENTADOS - DESVÍOS' in wb_dev.sheet_names:
                        df_dev_cm = wb_dev.parse('CM PRESENTADOS - DESVÍOS', skiprows=2)
                        df_dev_cm.columns = [str(c).strip() for c in df_dev_cm.columns]
                        col_desv  = next((c for c in df_dev_cm.columns if 'DESVÍO' in c or 'DESVIO' in c.upper()), None)
                        col_param = next((c for c in df_dev_cm.columns if 'PARÁMETRO' in c or 'PARAMETRO' in c.upper()), None)
                        col_exp   = next((c for c in df_dev_cm.columns if 'EXPEDIENTE' in c.upper() or 'EXP' in c.upper()), None)

                        if col_exp and col_param:
                            exp_map = {}
                            for _, row in df_dev_cm.iterrows():
                                exp = str(row.get(col_exp, '')).strip()
                                if exp:
                                    exp_map[exp] = {
                                        'desvio_desc': str(row.get(col_desv, '') or '').strip(),
                                        'parametro':   str(row.get(col_param,'') or '').strip()
                                    }
                            for item in cm_pre_items:
                                if str(item['exp']).strip() in exp_map:
                                    item['desvio_desc'] = exp_map[str(item['exp']).strip()]['desvio_desc']
                                    item['parametro']   = exp_map[str(item['exp']).strip()]['parametro']

                    # ── Verificar completitud ──
                    pendientes_lib = [i for i in lib_items  if i['desvio'] and not i.get('parametro','').strip()]
                    pendientes_ofi = [i for i in ofi_items  if i['desvio'] and not i.get('parametro','').strip()]
                    pendientes_cm  = [i for i in cm_pre_items if i['desvio'] and not i.get('parametro','').strip()]
                    total_pend = len(pendientes_lib) + len(pendientes_ofi) + len(pendientes_cm)

                    if total_pend == 0:
                        st.markdown('<div class="alert-success">✅ Todos los desvíos fueron completados correctamente.</div>', unsafe_allow_html=True)
                        st.markdown("<br>", unsafe_allow_html=True)

                        # Preview tabla
                        todos = (
                            [{'Proceso':'LIBERADAS', 'Ref': i['ref'], 'Razón':i['nombre'],
                              'Vía':i['via'], 'Canal':i.get('canal',''), 'Hs':i['hs'],
                              'Límite':i['limite'], 'Desvío':i['desvio_desc'], 'Parámetro':i['parametro']}
                             for i in lib_items if i['desvio']] +
                            [{'Proceso':'OFICIALIZADOS', 'Ref': i['ref'], 'Razón':i['nombre'],
                              'Vía':i['via'], 'Canal':'', 'Hs':i['hs'],
                              'Límite':i['limite'], 'Desvío':i['desvio_desc'], 'Parámetro':i['parametro']}
                             for i in ofi_items if i['desvio']] +
                            [{'Proceso':'CM PRES.', 'Ref': i['exp'], 'Razón':'',
                              'Vía':'', 'Canal':'', 'Hs':i['hs'],
                              'Límite':48, 'Desvío':i['desvio_desc'], 'Parámetro':i['parametro']}
                             for i in cm_pre_items if i['desvio']]
                        )
                        df_prev = pd.DataFrame(todos)
                        # Color por parámetro
                        def color_param(val):
                            if str(val).upper() == 'INTERLOG':
                                return 'background-color: rgba(255,61,94,0.2); color: #FF3D5E; font-weight:bold'
                            elif val:
                                return 'background-color: rgba(0,201,167,0.15); color: #00C9A7; font-weight:bold'
                            return ''
                        st.dataframe(
                            df_prev.style.map(color_param, subset=['Parámetro']),
                            use_container_width=True, hide_index=True
                        )

                        c1, c2 = st.columns([1, 3])
                        with c2:
                            if st.button("▶  GENERAR DASHBOARD", use_container_width=True):
                                st.session_state.lib_items    = lib_items
                                st.session_state.ofi_items    = ofi_items
                                st.session_state.cm_pre_items = cm_pre_items
                                st.session_state.desvio_sub_step = 'descargar'
                                st.session_state.step = 3
                                st.rerun()
                    else:
                        st.markdown(f'<div class="alert-warn">⚠️ Faltan completar {total_pend} desvíos. Revisá el Excel y volvé a subirlo.</div>', unsafe_allow_html=True)
                        if pendientes_lib:
                            st.markdown("**Sin completar en LIBERADAS:**")
                            for i in pendientes_lib:
                                st.markdown(f"- `{i['ref']}` · {i['nombre']} · {i['via']} {i.get('canal','')}")
                        if pendientes_ofi:
                            st.markdown("**Sin completar en OFICIALIZADOS:**")
                            for i in pendientes_ofi:
                                st.markdown(f"- `{i['ref']}` · {i['nombre']}")
                        if pendientes_cm:
                            st.markdown("**Sin completar en CM PRESENTADOS:**")
                            for i in pendientes_cm:
                                st.markdown(f"- `{i['exp']}`")

                except Exception as e:
                    st.markdown(f'<div class="alert-warn">❌ Error al leer el Excel: {str(e)}</div>', unsafe_allow_html=True)

            c1, c2 = st.columns([1, 3])
            with c1:
                if st.button("◀  VOLVER"):
                    st.session_state.desvio_sub_step = 'descargar'
                    st.rerun()

# ══════════════════════════════════════════════
# STEP 3 — DASHBOARD
# ══════════════════════════════════════════════
elif st.session_state.step == 3:
    lib_items    = st.session_state.lib_items
    ofi_items    = st.session_state.ofi_items
    cm_pre_items = st.session_state.cm_pre_items
    cm_apr_items = st.session_state.cm_apr_items
    mes          = st.session_state.mes or 'MENSUAL'

    st.markdown(f'<div class="section-header">📊 DASHBOARD KPI · {mes}</div>', unsafe_allow_html=True)

    # ── RESUMEN EJECUTIVO ──
    pct_lib_fasa, _, _ = calcular_kpi([i for i in lib_items if i['nombre'] == 'FASA'], True)
    pct_lib_fsm,  _, _ = calcular_kpi([i for i in lib_items if i['nombre'] == 'FSM'],  True)
    pct_ofi_fasa, _, _ = calcular_kpi([i for i in ofi_items if i['nombre'] == 'FASA'], True)
    pct_ofi_fsm,  _, _ = calcular_kpi([i for i in ofi_items if i['nombre'] == 'FSM'],  True)
    pct_cm, _, _       = calcular_kpi(cm_pre_items, True)

    def via_breakdown(items, nombre):
        """Genera HTML con desglose por vía para la card de liberación"""
        lines = []
        for via, emoji in [('AVION','✈️'), ('CAMION','🚛'), ('MARITIMO','🚢')]:
            sub = [i for i in items if i['nombre'] == nombre and i['via'] == via]
            if not sub: continue
            pct, _, _ = calcular_kpi(sub, True)
            color = '#00C9A7' if pct >= 95 else ('#FF8C42' if pct >= 80 else '#FF3D5E')
            lines.append(f'<span style="color:{color};font-size:0.75rem;font-weight:600;">'
                         f'{emoji} {via[:3]} {pct:.0f}%</span>')
        return '  ·  '.join(lines)

    c1, c2, c3, c4, c5 = st.columns(5)

    # LIB FASA con desglose
    col_lib_fasa = '#00C9A7' if pct_lib_fasa >= 95 else ('#FF8C42' if pct_lib_fasa >= 80 else '#FF3D5E')
    c1.markdown(f"""
    <div class="metric-card">
        <div class="metric-value" style="color:{col_lib_fasa}">{pct_lib_fasa:.0f}%</div>
        <div class="metric-label">LIB FASA</div>
        <div style="margin-top:0.4rem; line-height:1.8;">{via_breakdown(lib_items,'FASA')}</div>
    </div>""", unsafe_allow_html=True)

    # LIB FSM con desglose
    col_lib_fsm = '#00C9A7' if pct_lib_fsm >= 95 else ('#FF8C42' if pct_lib_fsm >= 80 else '#FF3D5E')
    c2.markdown(f"""
    <div class="metric-card">
        <div class="metric-value" style="color:{col_lib_fsm}">{pct_lib_fsm:.0f}%</div>
        <div class="metric-label">LIB FSM</div>
        <div style="margin-top:0.4rem; line-height:1.8;">{via_breakdown(lib_items,'FSM')}</div>
    </div>""", unsafe_allow_html=True)

    # OFI FASA, OFI FSM, CM — igual que antes
    for col, val, lbl, sub, pct in [
        (c3, f"{pct_ofi_fasa:.0f}%", "OFI FASA", "Oficialización", pct_ofi_fasa),
        (c4, f"{pct_ofi_fsm:.0f}%",  "OFI FSM",  "Oficialización", pct_ofi_fsm),
        (c5, f"{pct_cm:.0f}%",       "CM PRES.", "Certificados Mineros", pct_cm),
    ]:
        color = 'accent' if pct >= 95 else 'orange'
        col.markdown(metric_html(val, lbl, sub, color), unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── TABS ──
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["📦 FASA", "🏔️ FSM", "📊 DETALLE POR VÍA", "📊 DISTRIBUCIÓN DE CANALES", "📜 CERTIFICADOS MINEROS"])

    with tab1:
        render_kpi_section('FASA', lib_items, ofi_items, con_param=True)

    with tab2:
        render_kpi_section('FSM', lib_items, ofi_items, con_param=True)

    with tab3:
        st.markdown('<div class="section-header">DETALLE DE LIBERACIONES · POR SOCIEDAD, VÍA Y CANAL</div>', unsafe_allow_html=True)
        vias_orden = ['AVION', 'MARITIMO', 'CAMION']
        canales_orden = ['VERDE', 'NARANJA', 'ROJO']
        razones = ['FASA', 'FSM']

        for razon in razones:
            r_items = [i for i in lib_items if i['nombre'] == razon]
            if not r_items: continue
            accent_color = '#00C9A7' if razon == 'FASA' else '#FF8C42'
            st.markdown(f"""
            <div style="background:#132236; border-left:4px solid {accent_color};
                 padding:0.6rem 1rem; margin:1.2rem 0 0.6rem; border-radius:0 6px 6px 0;">
                <span style="font-family:'Barlow Condensed',sans-serif; font-size:1.2rem;
                font-weight:800; color:{accent_color}; letter-spacing:2px;">{razon}</span>
                <span style="font-size:0.8rem; color:#6B8099; margin-left:1rem;">
                {len(r_items)} liberaciones totales</span>
            </div>""", unsafe_allow_html=True)

            for via in vias_orden:
                via_items = [i for i in r_items if i['via'] == via]
                if not via_items: continue
                pct_via, in_via, out_via = calcular_kpi(via_items, True)
                kpi_color = '#00C9A7' if pct_via >= 95 else ('#FF8C42' if pct_via >= 80 else '#FF3D5E')
                st.markdown(f"""
                <div style="background:#1A2E48; border-radius:6px; padding:0.5rem 1rem; margin:0.4rem 0;">
                    <span style="font-family:'Barlow Condensed',sans-serif; font-weight:700; color:#F0F4F8; font-size:1rem; min-width:100px;">{'✈️' if via=='AVION' else '🚛' if via=='CAMION' else '🚢'} {via}</span>
                    <span style="color:{kpi_color}; font-weight:800; font-size:1.1rem; font-family:'Barlow Condensed',sans-serif; margin-left:1rem;">{pct_via:.1f}% IN</span>
                    <span style="color:#6B8099; font-size:0.8rem; margin-left:0.5rem;">| {len(via_items)} ops · {in_via} IN · {out_via} OUT</span>
                </div>""", unsafe_allow_html=True)
                cols_canal = st.columns(len(canales_orden))
                for col, canal in zip(cols_canal, canales_orden):
                    c_items = [i for i in via_items if i['canal'] == canal]
                    if not c_items:
                        col.markdown(f"""<div style="background:#0F2040; border-radius:6px; padding:0.8rem; text-align:center; opacity:0.3; border:1px dashed #1A2E48;">
                            <div style="color:#6B8099; font-size:0.75rem; font-weight:600;">CANAL {canal}</div>
                            <div style="color:#6B8099;">—</div></div>""", unsafe_allow_html=True)
                        continue
                    pct_c, in_c, out_c = calcular_kpi(c_items, True)
                    canal_colors = {'VERDE': '#00C9A7', 'NARANJA': '#FF8C42', 'ROJO': '#FF3D5E'}
                    cc = canal_colors.get(canal, '#6B8099')
                    kpi_c = '#00C9A7' if pct_c >= 95 else ('#FF8C42' if pct_c >= 80 else '#FF3D5E')
                    dias_vals = [i['hs'] for i in c_items if i['hs'] is not None]
                    avg_dias = f"{np.mean(dias_vals):.1f}d" if dias_vals else "—"
                    col.markdown(f"""<div style="background:#132236; border-top:3px solid {cc}; border-radius:6px; padding:0.8rem; text-align:center; border:1px solid rgba(255,255,255,0.05);">
                        <div style="color:{cc}; font-size:0.7rem; font-weight:700; letter-spacing:1px;">CANAL {canal}</div>
                        <div style="color:{kpi_c}; font-size:1.6rem; font-weight:800; font-family:'Barlow Condensed',sans-serif; margin:0.2rem 0;">{pct_c:.1f}%</div>
                        <div style="color:#9AB0C4; font-size:0.75rem;">{len(c_items)} ops</div>
                        <div style="display:flex; justify-content:center; gap:0.5rem; margin-top:0.3rem;">
                            <span style="color:#00C9A7; font-size:0.7rem; font-weight:600;">✓{in_c}</span>
                            <span style="color:#FF3D5E; font-size:0.7rem; font-weight:600;">✗{out_c}</span>
                        </div>
                        <div style="color:#6B8099; font-size:0.65rem; margin-top:0.2rem;">Prom: {avg_dias} · Lím: {c_items[0]['limite']}d</div>
                    </div>""", unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)

            with st.expander(f"📋 Tabla resumen {razon}"):
                filas = []
                for via in vias_orden:
                    for canal in canales_orden:
                        c_items = [i for i in r_items if i['via'] == via and i['canal'] == canal]
                        if not c_items: continue
                        pct_c, in_c, out_c = calcular_kpi(c_items, True)
                        dias_vals = [i['hs'] for i in c_items if i['hs'] is not None]
                        filas.append({'Vía': via, 'Canal': canal, 'Total': len(c_items),
                                      'IN': in_c, 'OUT': out_c, 'KPI %': f"{pct_c:.1f}%",
                                      'Prom días': f"{np.mean(dias_vals):.1f}" if dias_vals else '—',
                                      'Límite días': c_items[0]['limite']})
                if filas:
                    st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)
            st.markdown("---")

    with tab4:
        st.markdown('<div class="section-header">DISTRIBUCIÓN DE CANALES · POR VÍA Y SOCIEDAD</div>', unsafe_allow_html=True)

        # General — todas las razones sociales
        st.markdown("""<div style="color:#FFD060; font-family:'Barlow Condensed',sans-serif;
            font-size:1rem; font-weight:700; letter-spacing:1px; margin:0.5rem 0;">GENERAL</div>""", unsafe_allow_html=True)
        vias_orden2 = [('AVION','✈️'), ('CAMION','🚛'), ('MARITIMO','🚢')]
        cols_v = st.columns(len(vias_orden2))
        for col, (via, emoji) in zip(cols_v, vias_orden2):
            v_items = [i for i in lib_items if i['via'] == via]
            if not v_items:
                col.markdown(f'<div style="color:#6B8099; text-align:center;">{emoji} {via}<br>Sin ops</div>', unsafe_allow_html=True)
                continue
            by_canal = Counter(i['canal'] for i in v_items)
            fig = go.Figure(go.Pie(
                labels=list(by_canal.keys()), values=list(by_canal.values()),
                hole=0.55,
                marker_colors=[{'VERDE': COLORS['verde'], 'NARANJA': COLORS['naranja'], 'ROJO': COLORS['rojo']}.get(k, COLORS['gray']) for k in by_canal.keys()],
                textinfo='label+percent', textfont=dict(color='white', size=11)
            ))
            fig.update_layout(**CHART_LAYOUT, height=250, title=f"{emoji} {via} ({len(v_items)} ops)",
                showlegend=False)
            col.plotly_chart(fig, use_container_width=True, key=f"pie_gen_{via.lower()}")

        # Por sociedad
        for razon in ['FASA', 'FSM']:
            accent_color = '#00C9A7' if razon == 'FASA' else '#FF8C42'
            st.markdown(f"""<div style="color:{accent_color}; font-family:'Barlow Condensed',sans-serif;
                font-size:1rem; font-weight:700; letter-spacing:1px; margin:1rem 0 0.5rem;">{razon}</div>""", unsafe_allow_html=True)
            cols_v2 = st.columns(len(vias_orden2))
            for col, (via, emoji) in zip(cols_v2, vias_orden2):
                v_items = [i for i in lib_items if i['nombre'] == razon and i['via'] == via]
                if not v_items:
                    col.markdown(f'<div style="color:#6B8099; text-align:center; padding:1rem;">{emoji} Sin ops</div>', unsafe_allow_html=True)
                    continue
                by_canal = Counter(i['canal'] for i in v_items)
                fig = go.Figure(go.Pie(
                    labels=list(by_canal.keys()), values=list(by_canal.values()),
                    hole=0.55,
                    marker_colors=[{'VERDE': COLORS['verde'], 'NARANJA': COLORS['naranja'], 'ROJO': COLORS['rojo']}.get(k, COLORS['gray']) for k in by_canal.keys()],
                    textinfo='label+percent', textfont=dict(color='white', size=11)
                ))
                fig.update_layout(**CHART_LAYOUT, height=220, title=f"{emoji} {via} ({len(v_items)})",
                    showlegend=False)
                col.plotly_chart(fig, use_container_width=True, key=f"pie_{razon.lower()}_{via.lower()}")

    with tab5:
        st.markdown('<div class="section-header">KPI CERTIFICADOS MINEROS</div>', unsafe_allow_html=True)
        pct_cm_pre, in_cm, out_cm = calcular_kpi(cm_pre_items, True)

        c1,c2,c3,c4 = st.columns(4)
        c1.plotly_chart(chart_gauge(pct_cm_pre, "KPI Presentación"), use_container_width=True, key="gauge_cm")
        c2.markdown(metric_html(str(len(cm_pre_items)), "Total expedientes", "Presentados"), unsafe_allow_html=True)
        c3.markdown(metric_html(str(in_cm), "IN", "Dentro del rango", 'accent'), unsafe_allow_html=True)
        c4.markdown(metric_html(str(out_cm), "OUT", "Fuera del rango", 'red' if out_cm else 'accent'), unsafe_allow_html=True)

        # Desvíos en CM Presentados
        cm_dev = [i for i in cm_pre_items if i['desvio']]
        if cm_dev:
            st.markdown('<div class="section-header" style="border-color:#FF3D5E; margin-top:1rem;">DESVÍOS EN CM PRESENTADOS</div>', unsafe_allow_html=True)
            filas_cm = []
            for i in cm_dev:
                param = i.get('parametro', '')
                estado = '⛔ OUT' if param.upper() == 'INTERLOG' else ('✅ IN' if param else '⏳ Pendiente')
                filas_cm.append({'Expediente': i['exp'], 'Hs Transcurridas': i['hs'],
                                 'Límite (hs)': 48, 'Desvío': i.get('desvio_desc',''),
                                 'Parámetro': param, 'Estado': estado})
            st.dataframe(pd.DataFrame(filas_cm), use_container_width=True, hide_index=True)

        # CM Aprobados — donut por rango
        if cm_apr_items:
            st.markdown('<div class="section-header" style="margin-top:1rem;">CM APROBADOS · TIEMPOS DE APROBACIÓN</div>', unsafe_allow_html=True)
            rangos = Counter(i['rango'] for i in cm_apr_items if i['rango'])
            labels_r = ['0 a 7 días', '8 a 15 días', '+15 días']
            keys_r   = ['0 a 7', '8 a 15', '+15']
            values_r = [rangos.get(k, 0) for k in keys_r]
            total_apr = sum(values_r)

            c1, c2 = st.columns([1, 2])
            with c1:
                fig_donut = go.Figure(go.Pie(
                    labels=labels_r, values=values_r, hole=0.6,
                    marker_colors=[COLORS['verde'], COLORS['naranja'], COLORS['rojo']],
                    textinfo='none',
                    hovertemplate='%{label}: %{value} exp (%{percent})<extra></extra>'
                ))
                fig_donut.add_annotation(text=f"<b>{total_apr}</b>", x=0.5, y=0.58,
                    font=dict(size=28, color='#F0F4F8', family='Barlow Condensed'), showarrow=False)
                fig_donut.add_annotation(text="aprobados", x=0.5, y=0.38,
                    font=dict(size=11, color='#6B8099', family='Barlow'), showarrow=False)
                fig_donut.update_layout(**CHART_LAYOUT, height=260, showlegend=True,
                    legend=dict(orientation='h', yanchor='bottom', y=-0.2,
                                font=dict(size=10, color='#9AB0C4')))
                st.plotly_chart(fig_donut, use_container_width=True, key="donut_cm_apr")
            with c2:
                for lbl, key, val, col_r in zip(labels_r, keys_r, values_r,
                                                 [COLORS['verde'], COLORS['naranja'], COLORS['rojo']]):
                    pct_r = val/total_apr*100 if total_apr else 0
                    st.markdown(f"""
                    <div style="background:#132236; border-left:4px solid {col_r}; border-radius:0 6px 6px 0;
                         padding:0.6rem 1rem; margin:0.3rem 0; display:flex; justify-content:space-between; align-items:center;">
                        <span style="color:#F0F4F8; font-weight:600; font-size:0.9rem;">{lbl}</span>
                        <span style="color:{col_r}; font-family:'Barlow Condensed',sans-serif;
                        font-size:1.3rem; font-weight:800;">{val} <span style="font-size:0.8rem; color:#6B8099;">({pct_r:.0f}%)</span></span>
                    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    c1, c2 = st.columns([1, 3])
    with c1:
        if st.button("◀  VOLVER"):
            st.session_state.step = 2
            st.rerun()
    with c2:
        if st.button("▶  EXPORTAR"):
            st.session_state.step = 4
            st.rerun()

# ══════════════════════════════════════════════
# STEP 4 — EXPORTAR
# ══════════════════════════════════════════════
elif st.session_state.step == 4:
    st.markdown('<div class="section-header">PASO 4 · EXPORTAR</div>', unsafe_allow_html=True)

    lib_items    = st.session_state.lib_items
    ofi_items    = st.session_state.ofi_items
    cm_pre_items = st.session_state.cm_pre_items
    cm_apr_items = st.session_state.cm_apr_items
    mes          = st.session_state.mes or 'MES'

    st.markdown("""
    <div class="alert-info">
    Descargá el <b>Excel completo del dashboard</b> con toda la información que sostiene los KPIs,
    o el <b>PowerPoint</b> del informe listo para presentar.
    </div><br>
    """, unsafe_allow_html=True)

    c1, c2 = st.columns(2)

    with c1:
        st.markdown("### 📊 Excel Dashboard Completo")
        st.markdown("""
        Incluye 5 hojas con toda la información del dashboard:
        Resumen ejecutivo · Liberaciones · Oficializaciones · CM Presentados · CM Aprobados.
        Cada operación con desvío incluye su **DESVÍO** y **PARÁMETRO**.
        """)
        if st.button("⚙️  GENERAR EXCEL", use_container_width=True, key="btn_excel"):
            with st.spinner("Generando Excel..."):
                try:
                    excel_buf = export_dashboard_excel(lib_items, ofi_items, cm_pre_items, cm_apr_items, mes)
                    st.download_button(
                        label="⬇  DESCARGAR EXCEL DASHBOARD",
                        data=excel_buf,
                        file_name=f"DASHBOARD_KPI_{mes.replace(' ','_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    st.markdown('<div class="alert-success">✅ Excel generado — 5 hojas</div>', unsafe_allow_html=True)
                except Exception as e:
                    st.markdown(f'<div class="alert-warn">❌ Error: {str(e)}</div>', unsafe_allow_html=True)

    with c2:
        st.markdown("### 📑 PowerPoint")
        st.markdown("Informe completo con todos los gráficos, listo para presentar.")
        if st.button("⚙️  GENERAR POWERPOINT", use_container_width=True):
            with st.spinner("Generando PowerPoint... (~20 segundos)"):
                try:
                    from ppt_generator import generar_ppt
                    ppt_buf = generar_ppt(
                        lib_items, ofi_items, cm_pre_items, cm_apr_items,
                        mes=st.session_state.mes or 'MES'
                    )
                    st.download_button(
                        label="⬇  DESCARGAR POWERPOINT",
                        data=ppt_buf,
                        file_name=f"KPI_INTERLOG_{(st.session_state.mes or 'MES').replace(' ','_')}.pptx",
                        mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                        use_container_width=True
                    )
                    st.markdown('<div class="alert-success">✅ PowerPoint generado correctamente — 9 slides</div>', unsafe_allow_html=True)
                except Exception as e:
                    st.markdown(f'<div class="alert-warn">❌ Error al generar PowerPoint: {str(e)}</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("◀  VOLVER AL DASHBOARD"):
        st.session_state.step = 3
        st.rerun()
